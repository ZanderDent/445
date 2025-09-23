#!/usr/bin/env python3
# xcel_to_schedule_json.py
# Parse "BRIDGEWORKS COLLECTIVE-Project-Schedule.xlsx" => schedule.json

from __future__ import annotations
import argparse, json, re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, date
from openpyxl import load_workbook

# --------- Helpers ---------

CURRENCY_RE = re.compile(r"[^\d\.\-]")

def to_str(x: Any) -> str:
    if x is None: return ""
    return str(x).strip()

def clean_budget(x: Any) -> Optional[float]:
    s = to_str(x)
    if not s: return None
    s = CURRENCY_RE.sub("", s)  # remove $, commas, spaces
    if s == "": return None
    try:
        return float(s)
    except ValueError:
        return None

def is_date_cell(val: Any) -> bool:
    return isinstance(val, (datetime, date))

def iso_date(dt: date | datetime) -> str:
    if isinstance(dt, datetime):
        return dt.date().isoformat()
    return dt.isoformat()

def infer_year_for_mmdd(mm: int, dd: int) -> int:
    # Project spans Sept 2025 → Apr 2026 per your brief.
    # Infer year when the cell shows like "9/21" or "1/10".
    # Months Sep–Dec → 2025; Jan–Apr → 2026; else fallback 2025.
    if 9 <= mm <= 12: return 2025
    if 1 <= mm <= 4:  return 2026
    return 2025

def parse_date_maybe_year(x: Any) -> Optional[str]:
    """
    Accepts:
      - Excel date/datetime
      - 'MM/DD' or 'M/D'
      - 'YYYY-MM-DD' or 'MM/DD/YYYY' or 'Mon DD, YYYY' (anything datetime.fromisoformat can’t do we try fallback)
    Returns ISO 'YYYY-MM-DD' or None.
    """
    if x is None: return None
    if is_date_cell(x):
        return iso_date(x)
    s = to_str(x)
    if not s: return None

    # Try easy ISO first
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except Exception:
        pass

    # Try MM/DD[/YYYY] patterns
    m = re.match(r"^\s*(\d{1,2})\s*[/-]\s*(\d{1,2})(?:\s*[/-]\s*(\d{2,4}))?\s*$", s)
    if m:
        mm = int(m.group(1))
        dd = int(m.group(2))
        if m.group(3):
            yy = int(m.group(3))
            if yy < 100: yy += 2000
        else:
            yy = infer_year_for_mmdd(mm, dd)
        try:
            return date(yy, mm, dd).isoformat()
        except Exception:
            return None

    # Try very loose parse like "Friday, September 12, 2025"
    try:
        # Supports e.g. "September 12, 2025", "Fri, Sep 12, 2025"
        return datetime.strptime(s, "%A, %B %d, %Y").date().isoformat()
    except Exception:
        pass
    for fmt in ("%b %d, %Y", "%B %d, %Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            continue

    return None

def squash_newlines(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()

# --------- Core parsing ---------

def parse_project_schedule_sheet(ws) -> Dict[str, Any]:
    """
    Expected columns in the main table (observed from your paste):
      A: Main Phase
      B: SUB - Phase Title (task)
      C: Start Date
      D: End Date
      E: Duration (in days)
      F: Schedule
      G: Budget
      H: Resources
    We'll detect the header row by looking for the literal 'Timeline' + 'Start Date' etc nearby.
    """
    # Scan entire sheet to collect meta and detect header row
    project_name = None
    project_manager = None
    project_start_label_row = None
    project_end_label_row = None
    header_row_idx: Optional[int] = None

    # Find metadata and header
    for r in ws.iter_rows():
        for c in r:
            val = to_str(c.value)
            if not val:
                continue
            u = val.upper()

            if "PROJECT NAME" in u and not project_name:
                # The actual name likely to the right or below
                # try same row, next non-empty cell
                project_name = find_neighbor_value(ws, c, prefer_row=True)

            if "PROJECT MANAGER" in u and not project_manager:
                project_manager = find_neighbor_value(ws, c, prefer_row=True)

            if val.strip() == "Timeline":
                # Likely header is on this row with more fields (Start Date, End Date, etc.) or the next row.
                header_row_idx = c.row

            if "START DATE" in u:
                project_start_label_row = c.row
            if "END DATE" in u:
                project_end_label_row = c.row

    # Determine project-level start/end from labeled rows (neighbor cell to right),
    # otherwise compute from tasks later.
    project_start = None
    project_end = None
    if project_start_label_row:
        for cell in ws[project_start_label_row]:
            if "START DATE" in to_str(cell.value).upper():
                project_start = parse_date_maybe_year(find_neighbor_value(ws, cell, prefer_row=True))
                break
    if project_end_label_row:
        for cell in ws[project_end_label_row]:
            if "END DATE" in to_str(cell.value).upper():
                project_end = parse_date_maybe_year(find_neighbor_value(ws, cell, prefer_row=True))
                break

    # If header_row_idx not found, try to guess: find row that contains "Start Date" and "End Date"
    if header_row_idx is None:
        for r in ws.iter_rows():
            texts = [to_str(c.value).lower() for c in r if to_str(c.value)]
            if ("start date" in texts) and ("end date" in texts):
                header_row_idx = r[0].row
                break

    # Fallback if still None
    if header_row_idx is None:
        # Heuristic: use first row where column A looks like a phase label and column B has text
        for r in ws.iter_rows(min_row=1, max_col=2):
            a, b = r[0], r[1]
            if to_str(a.value) and to_str(b.value):
                header_row_idx = r[0].row
                break

    # Now parse table rows under header
    tasks: List[Dict[str, Any]] = []
    current_phase = None

    # Determine column indexes by scanning header row for keywords
    # Default to A..H if not detected
    col_idx = {"phase": 1, "task": 2, "start": 3, "end": 4, "duration": 5, "schedule": 6, "budget": 7, "resources": 8}
    if header_row_idx is not None:
        header_cells = list(ws[header_row_idx])
        for j, cell in enumerate(header_cells, start=1):
            txt = to_str(cell.value).lower()
            if "start" in txt and "date" in txt: col_idx["start"] = j
            elif ("end" in txt and "date" in txt) or "finish" in txt: col_idx["end"] = j
            elif "duration" in txt: col_idx["duration"] = j
            elif "schedule" in txt: col_idx["schedule"] = j
            elif "budget" in txt: col_idx["budget"] = j
            elif "resource" in txt: col_idx["resources"] = j
            elif "sub" in txt or "phase title" in txt or "task" in txt: col_idx["task"] = j
            elif "main" in txt or "phase" in txt: col_idx["phase"] = j

    start_row = (header_row_idx or 1) + 1
    last_row = ws.max_row

    for r in range(start_row, last_row + 1):
        # Access by 1-based column index
        def val(ci: int) -> Any:
            return ws.cell(r, ci).value if 1 <= ci <= ws.max_column else None

        a = val(col_idx["phase"])
        b = val(col_idx["task"])
        c = val(col_idx["start"])
        d = val(col_idx["end"])
        e = val(col_idx["duration"])
        f = val(col_idx["schedule"])
        g = val(col_idx["budget"])
        h = val(col_idx["resources"])

        # Detect phase rows (main phase in col A; task is empty)
        if to_str(a) and not to_str(b):
            current_phase = squash_newlines(to_str(a).replace('"', ""))
            continue

        # Skip empty rows (no task name)
        if not to_str(b):
            continue

        task_name = squash_newlines(to_str(b))

        start_iso = parse_date_maybe_year(c)
        end_iso   = parse_date_maybe_year(d)

        # Duration: prefer explicit; else compute if start/end available (inclusive)
        duration = None
        if to_str(e):
            try:
                duration = int(float(to_str(e)))
            except ValueError:
                duration = None
        if duration is None and start_iso and end_iso:
            d0 = datetime.fromisoformat(start_iso).date()
            d1 = datetime.fromisoformat(end_iso).date()
            duration = (d1 - d0).days + 1

        schedule = squash_newlines(to_str(f)) if to_str(f) else None
        budget   = clean_budget(g)
        resources= squash_newlines(to_str(h)) if to_str(h) else None

        tasks.append({
            "phase": current_phase or "",
            "task": task_name,
            "start": start_iso,
            "finish": end_iso,
            "duration": duration,
            "budget": budget,
            "schedule": schedule,
            "resources": resources
        })

    # If project_name/manager missing, try to infer from top-left cells
    if not project_name:
        project_name = guess_near(ws, "Project Name")
    if not project_manager:
        project_manager = guess_near(ws, "Project Manager")

    # Derive project start/end if not captured
    if not project_start or not project_end:
        starts = [t["start"] for t in tasks if t["start"]]
        ends   = [t["finish"] for t in tasks if t["finish"]]
        if starts and not project_start: project_start = min(starts)
        if ends and not project_end:     project_end   = max(ends)

    return {
        "project": project_name or "TSOLOM BRIDGE PROJECT",
        "manager": project_manager or "BRIDGEWORKS COLLECTIVE",
        "start": project_start or "2025-09-12",
        "finish": project_end or "2026-04-22",
        "tasks": tasks
    }

def find_neighbor_value(ws, cell, prefer_row=True) -> Any:
    """Find next non-empty neighbor to the right (prefer_row=True) else below."""
    if prefer_row:
        col = cell.column + 1
        while col <= ws.max_column:
            v = ws.cell(cell.row, col).value
            if to_str(v):
                return v
            col += 1
    else:
        row = cell.row + 1
        while row <= ws.max_row:
            v = ws.cell(row, cell.column).value
            if to_str(v):
                return v
            row += 1
    return None

def guess_near(ws, label: str) -> Optional[str]:
    L = label.upper()
    for r in ws.iter_rows():
        for c in r:
            if L in to_str(c.value).upper():
                v = find_neighbor_value(ws, c, prefer_row=True)
                return to_str(v) if v else None
    return None

# --------- Entry ---------

def convert_excel_to_json(in_path: Path, out_path: Path, sheet_name: str = "PROJECT SCHEDULE ") -> None:
    wb = load_workbook(in_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        # fallback: find first sheet containing "SCHEDULE"
        candidates = [s for s in wb.sheetnames if "SCHEDULE" in s.upper()]
        if not candidates:
            raise ValueError(f"Could not find a schedule sheet in: {wb.sheetnames}")
        sheet_name = candidates[0]
    ws = wb[sheet_name]
    payload = parse_project_schedule_sheet(ws)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {out_path} with {len(payload.get('tasks', []))} tasks")

def main():
    p = argparse.ArgumentParser(description="Convert Bridgeworks project schedule Excel → JSON.")
    p.add_argument("excel", nargs="?", default="BRIDGEWORKS COLLECTIVE-Project-Schedule.xlsx")
    p.add_argument("out", nargs="?", default="static/schedule.json")
    p.add_argument("--sheet", default="PROJECT SCHEDULE ", help="Sheet name (default exact match).")
    args = p.parse_args()
    convert_excel_to_json(Path(args.excel), Path(args.out), sheet_name=args.sheet)

if __name__ == "__main__":
    main()
