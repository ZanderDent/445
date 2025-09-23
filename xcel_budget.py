#!/usr/bin/env python3
# xcel_to_schedule_json.py
# Parse "BRIDGEWORKS COLLECTIVE-Project-Schedule.xlsx" => schedule.json

from __future__ import annotations
import argparse, json, re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, date
from openpyxl import load_workbook

# --------- Helpers ---------

CURRENCY_RE = re.compile(r"[^\d\.\-]")

def to_str(x: Any) -> str:
    if x is None: return ""
    return str(x).strip()

def clean_budget(x: Any) -> Optional[float]:
    s = to_str(x)
    if not s: return None
    s = CURRENCY_RE.sub("", s)  # remove $, commas, spaces
    if s == "": return None
    try:
        return float(s)
    except ValueError:
        return None

def is_date_cell(val: Any) -> bool:
    return isinstance(val, (datetime, date))

def iso_date(dt: date | datetime) -> str:
    if isinstance(dt, datetime):
        return dt.date().isoformat()
    return dt.isoformat()

def infer_year_for_mmdd(mm: int, dd: int) -> int:
    # Project spans Sept 2025 → Apr 2026 per your brief.
    # Infer year when the cell shows like "9/21" or "1/10".
    # Months Sep–Dec → 2025; Jan–Apr → 2026; else fallback 2025.
    if 9 <= mm <= 12: return 2025
    if 1 <= mm <= 4:  return 2026
    return 2025

def parse_date_maybe_year(x: Any) -> Optional[str]:
    """
    Accepts:
      - Excel date/datetime
      - 'MM/DD' or 'M/D'
      - 'YYYY-MM-DD' or 'MM/DD/YYYY' or 'Mon DD, YYYY' (anything datetime.fromisoformat can’t do we try fallback)
    Returns ISO 'YYYY-MM-DD' or None.
    """
    if x is None: return None
    if is_date_cell(x):
        return iso_date(x)
    s = to_str(x)
    if not s: return None

    # Try easy ISO first
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except Exception:
        pass

    # Try MM/DD[/YYYY] patterns
    m = re.match(r"^\s*(\d{1,2})\s*[/-]\s*(\d{1,2})(?:\s*[/-]\s*(\d{2,4}))?\s*$", s)
    if m:
        mm = int(m.group(1))
        dd = int(m.group(2))
        if m.group(3):
            yy = int(m.group(3))
            if yy < 100: yy += 2000
        else:
            yy = infer_year_for_mmdd(mm, dd)
        try:
            return date(yy, mm, dd).isoformat()
        except Exception:
            return None

    # Try very loose parse like "Friday, September 12, 2025"
    try:
        # Supports e.g. "September 12, 2025", "Fri, Sep 12, 2025"
        return datetime.strptime(s, "%A, %B %d, %Y").date().isoformat()
    except Exception:
        pass
    for fmt in ("%b %d, %Y", "%B %d, %Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            continue

    return None

def squash_newlines(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()

# --------- Core parsing ---------

def parse_project_schedule_sheet(ws) -> Dict[str, Any]:
    # If header_row_idx not found, try to guess: find row that contains "Start Date" and "End Date"
    if header_row_idx is None:
        for r in ws.iter_rows():
            texts = [to_str(c.value).lower() for c in r if to_str(c.value)]
            if ("start date" in texts) and ("end date" in texts):
                header_row_idx = r[0].row
                break

    # Fallback if still None
    if header_row_idx is None:
        # Heuristic: use first row where column A looks like a phase label and column B has text
        for r in ws.iter_rows(min_row=1, max_col=2):
            a, b = r[0], r[1]
            if to_str(a.value) and to_str(b.value):
                header_row_idx = r[0].row
                break

    # Now parse table rows under header
    tasks: List[Dict[str, Any]] = []
    current_phase = None

    # Determine column indexes by scanning header row for keywords
    # Default to A..H if not detected
    col_idx = {"role": 1, "hours": 2, "rate": 3, "cost": 4}
    if header_row_idx is not None:
        header_cells = list(ws[header_row_idx])
        for j, cell in enumerate(header_cells, start=1):
            txt = to_str(cell.value).lower()

    start_row = (header_row_idx or 1) + 1
    last_row = ws.max_row

    for r in range(start_row, last_row + 1):
        # Access by 1-based column index
        def val(ci: int) -> Any:
            return ws.cell(r, ci).value if 1 <= ci <= ws.max_column else None

        a = val(col_idx["role"])
        b = val(col_idx["hours"])
        c = val(col_idx["rate"])
        d = val(col_idx["cost"])


def convert_excel_to_json(in_path: Path, out_path: Path, sheet_name: str = "PROJECT SCHEDULE ") -> None:
    wb = load_workbook(in_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        # fallback: find first sheet containing "SCHEDULE"
        candidates = [s for s in wb.sheetnames if "SCHEDULE" in s.upper()]
        if not candidates:
            raise ValueError(f"Could not find a schedule sheet in: {wb.sheetnames}")
        sheet_name = candidates[0]
    ws = wb[sheet_name]
    payload = parse_project_schedule_sheet(ws)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {out_path} with {len(payload.get('tasks', []))} tasks")

def main():
    p = argparse.ArgumentParser(description="Convert Bridgeworks project schedule Excel → JSON.")
    p.add_argument("excel", nargs="?", default="BRIDGEWORKS COLLECTIVE-Project-Schedule.xlsx")
    p.add_argument("out", nargs="?", default="static/schedule.json")
    p.add_argument("--sheet", default="PROJECT SCHEDULE ", help="Sheet name (default exact match).")
    args = p.parse_args()
    convert_excel_to_json(Path(args.excel), Path(args.out), sheet_name=args.sheet)

if __name__ == "__main__":
    main()
